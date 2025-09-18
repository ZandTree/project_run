from django.conf import settings
from django.contrib.auth.models import User
from django.db.models import Avg, Count, Max, Q
from django.db.models.functions import Round
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import api_view
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.response import Response
from rest_framework.views import APIView

from app_run.models import (AthleteInfo, Challenge, CollectibleItem, Position,
                            Run, Subscribe)

from .pagination import CustomPagination
from .serializers import (AthleteInfoSerializer, ChallengeSerializer,
                          ChallengeSummarySerializer,
                          CollectibleItemSerializer, PositionSerializer,
                          RunSerializer, SubscribeSerializer,
                          UserAthleteSerializer, UserCoachSerializer,
                          UserSerializer)
from .utils import (calc_total_distance, create_dict, get_total_distance,
                    get_total_time, parse_positions)


@api_view(['GET'])
def get_intro(request):    
    company_dict = {
        "company_name":settings.COMPANY_NAME,
        "slogan":settings.SLOGAN,
        "contacts":settings.CONTACTS
    }    
    return Response(company_dict)

class RunViewSetClass(viewsets.ModelViewSet):    
    """
    api/runs/...
    """   
    queryset = Run.objects.select_related('athlete').all()    
    serializer_class = RunSerializer 
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend,OrderingFilter]    
    ordering_fields = ['created_at',]
    filterset_fields = ['status','athlete']

      
class FilterAthletesClass(viewsets.ReadOnlyModelViewSet): 
    """
    all coaches and athletes (no su): order, filter, search, pagination
    /api/users/
    "search" default; change to "q" via settings in drf dict SEARCH_PARAM     
    """  
    serializer_class = UserSerializer              
    queryset = User.objects.prefetch_related('items').filter(is_superuser=False).annotate(count=Count('runs',filter=Q(runs__status="finished")))       
    filter_backends = [SearchFilter,OrderingFilter]
    search_fields = ['first_name', 'last_name','date_joined']
    pagination_class = CustomPagination

    
    def get_queryset(self):
        """
        users selected via url query params: 
        qs only coaches vs qs only athlete
        """
        # qs = self.queryset 
        qs = self.queryset.annotate(rating=Round((Avg('coaches__rating')),2))                        
        type = self.request.query_params.get("type",None) 
        if type:
            if type == "coach":
                return qs.filter(is_staff=True)
            elif type == "athlete":
                return qs.filter(is_staff=False)
        return qs  
        
    def get_serializer_class(self):
        """
        provide diff serializers;
        detail(retrieve): return detailed ser-er (coaches vs athletes)
        """
        if self.action == "list":            
            return UserSerializer
        elif self.action == "retrieve":             
            # self.kwargs = {"pk":1}          
            user = self.get_object()            
            if user.is_staff:
                return UserCoachSerializer
            else:
                return UserAthleteSerializer
        return super().get_serializer_class()
    

class RunStart(APIView):
    """
    initate start run
    """
    def post(self, request, run_id, format=None):               
        try:
            run = get_object_or_404(Run, id=run_id)  
            if run.status == "init":           
                run.status = "in_progress"
                run.save()
                data = {"status":run.status}
                return Response(data,status=status.HTTP_200_OK)  
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)         
        except Run.DoesNotExist:            
            return Response(status=status.HTTP_404_NOT_FOUND)
    
class RunStop (APIView):
    """
    stopping runs/<run_id>/stop/
    """
    def post(self, request, run_id, format=None):
        """
        1. assigns run status to "finished"
        2. create Challenge Model objects:
        2.1 "Сделай 10 Забегов!";
        2.2 "Пробеги 50 километров!";
        2.3  "2 километра за 10 минут!"
        3. if geo data present |=> calc distance (float) of a given Run obj
        """
        try:            
            run = get_object_or_404(Run, id=run_id)
            if run.status == "in_progress":
                run.status = "finished"
                run.save()
                user = run.athlete 
                # create first challenge               
                if user.runs.filter(status="finished").count()%10 == 0:
                    Challenge.objects.create(
                        athlete = run.athlete, full_name = "Сделай 10 Забегов!"
                    )                
                run_positions = run.positions.all()   
                #           
                run.distance = get_total_distance(run_positions) 
                run.run_time_seconds = get_total_time(run_positions)        
                
                avg_speed = run_positions.aggregate(speed=Avg('speed'))                
                avg_speed = avg_speed.get("speed",0)                
                if not avg_speed:
                    avg_speed = 0                  
                run.speed = round(avg_speed,2)                                
                run.save()
                data = {"status":run.status}
                total = calc_total_distance(run)
                # create second challenge
                if  total >= 50:                    
                    Challenge.objects.create(full_name="Пробеги 50 километров!",athlete=run.athlete)
                # third challenge                  
                if run.distance >= 2 and run.run_time_seconds/60 <= 10:                                                 
                    Challenge.objects.create(full_name="2 километра за 10 минут!",athlete=run.athlete)
                return Response(data,status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)        
        except Run.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

class CreateUpdatePersonalInfo(APIView):
    def get(self,request,user_id):        
        user = get_object_or_404(User,id=user_id)        
        obj,created = AthleteInfo.objects.get_or_create(
            user=user) 
        data = AthleteInfoSerializer(obj).data                       
        return Response(data,status=status.HTTP_200_OK)
    
    def put(self,request,user_id):           
        user = get_object_or_404(User,id=user_id)         
        if not user:            
            return Response(status=status.HTTP_404_NOT_FOUND)        
        object,created = AthleteInfo.objects.update_or_create(
                user=user
            )        
        ser = AthleteInfoSerializer(object,data=request.data)  
        if ser.is_valid(raise_exception=True):   
            ser.save()
            data = ser.data
            if created:
                return Response(data=data,status=status.HTTP_201_CREATED)   
            else:
                return Response(data=data,status=status.HTTP_200_OK)

        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)
         

@api_view(['GET'])
def showChallenges(request):
    athlete_id = request.query_params.get("athlete",None)      
    if athlete_id:
        athlete = get_object_or_404(User,id=athlete_id)
        lst = athlete.challenges.all()       
    else:
        lst = Challenge.objects.all()
    data = ChallengeSerializer(lst,many=True).data
    return Response(data=data,status=status.HTTP_200_OK)


class PositionViewSet(viewsets.ModelViewSet):
    serializer_class = PositionSerializer        
    queryset = Position.objects.select_related("run").all()
    def get_queryset(self):   
        qs = self.queryset               
        run_id = self.request.query_params.get("run",None) 
        if run_id:            
            run_obj = get_object_or_404(Run,id=run_id)
            qs = run_obj.positions.all()             
        return qs           
         
    def perform_create(self,serializer):
        run = serializer.validated_data["run"]        
        run = get_object_or_404(Run,id=run.id)          
        prev_position =run.positions.last()   
        if prev_position:  
            # print("has prev position ")     
            super().perform_create(serializer)            
            next_position  = run.positions.last()
            speed,_distance = parse_positions(prev_position,next_position) 
            # print("speed is ", speed)           
            # print("distance is ", _distance)           
            next_position.speed = speed
            next_position.distance = _distance
            next_position.save()
            serializer.save(speed=speed,distance=_distance)
        else:            
            super().perform_create(serializer)


class ShowCollectibleItemSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = CollectibleItemSerializer       
    queryset = CollectibleItem.objects.all()    


@api_view(['POST'])
def create_items(request):
    """
    validating length of each row;
    creating dict data: zipping  model_fields(keys) and row (values)
    """    
    file_items = request.FILES.get("file")
    if file_items:
        wb = load_workbook(filename=file_items) 
        ctx = wb.active
        
        model_fields = ('name', 'uid', 'value', 'latitude', 'longitude', 'picture')             
        invalid_rows = []
        for row in ctx.iter_rows(min_row=2,values_only=True):
            if row[0]:
                data = create_dict(model_fields,row)                                                           
                row_ser = CollectibleItemSerializer(data=data)
                if row_ser.is_valid():
                    row_ser.save()                    
                else:                    
                    invalid_rows.append(list(row))
          
        if invalid_rows:
            return Response(data=invalid_rows)
        else:
            return Response(status=status.HTTP_201_CREATED)
        
    else:
        return Response(status=status.HTTP_400_BAD_REQUEST) 

   
class SubscribeView(APIView):
    def post(self,request,id):  
        """
        coach_id (subscribe_to_coach/<int:id>) and athlete id(body)
        to create a new subscription       
        """
        coach = get_object_or_404(User,id = id) 
        try:
            runner = User.objects.get(id = request.data["athlete"])          
        except User.DoesNotExist:
            return Response(status=status.HTTP_400_BAD_REQUEST)             
        inp = {"coach":coach.id,"runner":runner.id}        
        ser = SubscribeSerializer(data=inp)
        if ser.is_valid(raise_exception=True):
            ser.save()
            data = ser.data        
            return Response(data=data,status=status.HTTP_200_OK)
        else:            
            return Response(status=status.HTTP_400_BAD_REQUEST)



@api_view(['GET'])
def showChallenges(request):
    athlete_id = request.query_params.get("athlete",None)      
    if athlete_id:
        athlete = get_object_or_404(User,id=athlete_id)
        lst = athlete.challenges.all()       
    else:
        lst = Challenge.objects.all()
    data = ChallengeSerializer(lst,many=True).data
    return Response(data=data,status=status.HTTP_200_OK)


@api_view(['GET'])              
def summ_challenges(request): 
    """
    summarize challenge data
    """   
    ids = (Challenge.objects.values("full_name").annotate(max_id=Max("id"))        
       .values_list("max_id", flat=True))   
    challenges = Challenge.objects.filter(id__in=ids).select_related("athlete")               
    data = ChallengeSummarySerializer(challenges,many=True).data     
    return Response(data=data,status=status.HTTP_200_OK)

    
@api_view(['POST'])
def give_rating(request,coach_id):
    """
    rate_coach/<int:coach_id>/
    athlete id and rating are in request body
    """    
    athlete_id = get_object_or_404(User,id = request.data.get("athlete")) 
    rating = request.data.get("rating")
    coach = get_object_or_404(User,id=coach_id)
    if not rating or (rating < 1 or rating >5):        
        return Response(status=status.HTTP_400_BAD_REQUEST)
    try:
        subscribe  = get_object_or_404(Subscribe,coach_id=coach.id,runner_id=athlete_id) 
        if subscribe:
            subscribe.rating = rating
            subscribe.save()
        
    except Subscribe.DoesNotExist:
            return Response(status=status.HTTP_400_BAD_REQUEST)
            
    return Response(status=status.HTTP_200_OK)
    
    
    



